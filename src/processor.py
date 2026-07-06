import re
from pathlib import Path
from typing import Iterable

import pandas as pd
import pdfplumber
from openpyxl import load_workbook

SOCIETY_NAMES = {
    "PPAG": "PPAG",
    "PCP": "PCP - PHILIPPINE COLLEGE OF PHYSICIANS",
    "PCS": "PCS - PHILIPPINE COLLEGE OF SURGEONS",
    "POGS": "POGS - PHILIPPINE OBSTETRICAL AND GYNECOLOGICAL SOCIETY",
    "PAFP": "PAFP - PHILIPPINE ACADEMY OF FAMILY PHYSICIANS",
}

FINAL_COLUMNS = [
    "Doctor Name",
    "Medical Society",
    "Category",
    "Specialization",
    "Status",
    "Source File",
]


def detect_society(path: Path) -> str:
    name = path.name.upper()
    for key in ["PAFP", "PCP", "PCS", "POGS", "PPAG"]:
        if key in name:
            return key
    return "UNKNOWN"


def extract_text(pdf_path: Path) -> str:
    text_parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            text_parts.append(text)
    return "\n".join(text_parts)


def clean_name(value: str) -> str:
    value = re.sub(r"\s+", " ", value or "").strip()
    value = re.sub(r"\bMD\b\.?", "", value, flags=re.I).strip(" ,")
    return value


def blank_record(name: str, society: str, source: Path, category: str = "", specialization: str = "", status: str = "") -> dict:
    return {
        "Doctor Name": clean_name(name),
        "Medical Society": SOCIETY_NAMES.get(society, society),
        "Category": category.strip(),
        "Specialization": specialization.strip(),
        "Status": status.strip(),
        "Source File": source.name,
    }


def parse_numbered_names(text: str, society: str, source: Path) -> list[dict]:
    records = []
    for line in text.splitlines():
        line = re.sub(r"\s+", " ", line).strip()
        match = re.match(r"^(\d+)[\.)]?\s+(.+)$", line)
        if not match:
            continue
        payload = match.group(2).strip()
        if not payload or len(payload) < 4:
            continue
        payload = re.sub(r"\s+(FELLOW|DIPLOMATE|MEMBER)$", "", payload, flags=re.I)
        if any(word in payload.upper() for word in ["PAGE ", "DATE RECEIVED", "MEDICAL SOCIETY", "SUMMARY"]):
            continue
        records.append(blank_record(payload, society, source))
    return records


def parse_ppag_or_pafp(text: str, society: str, source: Path) -> list[dict]:
    records = []
    target_societies = ["PAFP", "PCP", "PCS", "POGS", "PPS", "PSA"] if society == "PPAG" else [society]
    pattern = re.compile(
        r"(?:^|\n)\s*\d+\s+\d{1,2}/\d{1,2}/\d{4}[^A-Z]*(?P<name>[A-ZÑ.\- ']+,\s*[A-ZÑ.\- ']+?)\s+"
        r"(?P<spec>FAMILY MEDICINE|INTERNAL MEDICINE|GENERAL SURGERY|ORTHOPEDIC SURGERY|OBSTETRICS AND GYNECOLOGY|PEDIATRICS|ANESTHESIOLOGY|DERMATOLOGY|UROLOGY|OPHTHALMOLOGY|OTOLARYNGOLOGY-HEAD AND NECK SURGERY)",
        re.I,
    )
    for match in pattern.finditer(text):
        raw_line = match.group(0)
        selected = society
        for code in target_societies:
            if code in raw_line.upper():
                selected = code
                break
        records.append(blank_record(match.group("name"), selected, source, specialization=match.group("spec")))
    if not records:
        records = parse_numbered_names(text, society, source)
    return records


def parse_pdf(pdf_path: Path) -> list[dict]:
    society = detect_society(pdf_path)
    text = extract_text(pdf_path)
    if society in {"PPAG", "PAFP"}:
        records = parse_ppag_or_pafp(text, society, pdf_path)
    else:
        records = parse_numbered_names(text, society, pdf_path)
    return [r for r in records if r["Doctor Name"]]


def remove_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    df["_key"] = df["Doctor Name"].str.upper().str.replace(r"[^A-ZÑ]", "", regex=True)
    df = df.drop_duplicates(subset=["_key", "Medical Society"], keep="first")
    return df.drop(columns=["_key"])


def write_to_template(df: pd.DataFrame, template_path: Path, output_path: Path) -> None:
    wb = load_workbook(template_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if not any(headers):
        headers = FINAL_COLUMNS
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_index, value=header)

    header_map = {str(header).strip(): index + 1 for index, header in enumerate(headers) if header}

    row_index = 2
    for _, row in df.iterrows():
        for column in FINAL_COLUMNS:
            excel_col = header_map.get(column)
            if excel_col:
                ws.cell(row=row_index, column=excel_col, value=row.get(column, ""))
        row_index += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def consolidate_pdfs_to_template(pdf_paths: Iterable[Path], template_path: Path, output_path: Path) -> dict:
    warnings = []
    all_records = []

    for pdf_path in pdf_paths:
        try:
            records = parse_pdf(Path(pdf_path))
            if not records:
                warnings.append(f"No records extracted from {Path(pdf_path).name}")
            all_records.extend(records)
        except Exception as exc:
            warnings.append(f"Failed to process {Path(pdf_path).name}: {exc}")

    df = pd.DataFrame(all_records, columns=FINAL_COLUMNS)
    df = remove_duplicates(df)
    write_to_template(df, Path(template_path), Path(output_path))

    return {"total_records": len(df), "warnings": warnings}
